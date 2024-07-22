import tkinter as tk
from tkinter import ttk
from tkcalendar import DateEntry
import openpyxl

def insert_row():
    Marque = Marque_entry.get()
    n_entry = N_entry.get()
    km = int(km_entry.get()) 
    date = cal.get()
    reparation = reparation_entry.get()
    qte = int(qte_spinbox.get())  
    prix = float(Prix_entry.get())  
    facture = facture_entry.get()
    ncarnet = Ncarnet_entry.get()

    # Insert row into Excel sheet starting from row 15
    path = r"C:\Users\VOSTRO 3500\OneDrive\Bureau\app\depense.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    
    # Find the next available row starting from row 15
    start_row = 15
    while sheet.cell(row=start_row, column=1).value is not None:
        start_row += 1
    
    row_values = [Marque, n_entry, km, date, reparation, qte, prix, facture, ncarnet]
    
    for col_num, value in enumerate(row_values, start=1):
        sheet.cell(row=start_row, column=col_num, value=value)
    
    workbook.save(path)

    # Clear the values
    Marque_entry.delete(0, "end")
    Marque_entry.insert(0, "Marque de Véhicule")
    N_entry.delete(0, "end")
    N_entry.insert(0, "N° d'immatriculation")
    km_entry.delete(0, "end")
    km_entry.insert(0, "Kilométrage")
    reparation_entry.delete(0, "end")
    reparation_entry.insert(0, "Réparation et pieces de rechange")
    qte_spinbox.delete(0, "end")
    qte_spinbox.insert(0, "Qté")
    Prix_entry.delete(0, "end")
    Prix_entry.insert(0, "Prix / unité H.T")
    facture_entry.delete(0, "end")
    facture_entry.insert(0, "N° de facture")
    Ncarnet_entry.delete(0, "end")
    Ncarnet_entry.insert(0, "N° Carnert vignette")

def toggle_mode():
    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")

root = tk.Tk()

style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

frame = ttk.Frame(root)
frame.grid(row=0, column=0, padx=20, pady=10)

widget_frame = ttk.LabelFrame(frame, text="Insert Row")
widget_frame.grid(row=0, column=0, padx=20, pady=10)

# Marque de Véhicule
Marque_entry = ttk.Entry(widget_frame)
Marque_entry.insert(0, "Marque de Véhicule")
Marque_entry.bind("<FocusIn>", lambda e: Marque_entry.delete('0', 'end'))
Marque_entry.grid(row=0, column=0, padx=20, pady=10, sticky="ew")

# N° d'immatriculation
N_entry = ttk.Entry(widget_frame)
N_entry.insert(0, "N° d'immatriculation")
N_entry.bind("<FocusIn>", lambda e: N_entry.delete('0', 'end'))
N_entry.grid(row=1, column=0, padx=20, pady=10, sticky="ew")

# Kilométrage
km_entry = ttk.Entry(widget_frame)
km_entry.insert(0, "Kilométrage")
km_entry.bind("<FocusIn>", lambda e: km_entry.delete('0', 'end'))
km_entry.grid(row=2, column=0, padx=20, pady=10, sticky="ew")

# Date
cal = DateEntry(widget_frame, width=30, year=2024, date_pattern='dd/mm/yyyy')
cal.grid(row=3, column=0, padx=20, pady=10, sticky="ew")

widget2_frame = ttk.LabelFrame(frame, text="Insert Row")
widget2_frame.grid(row=0, column=1, padx=10, pady=10)

# Réparation et pieces de rechange
reparation_entry = ttk.Entry(widget2_frame)
reparation_entry.insert(0, "Réparation et pieces de rechange")
reparation_entry.bind("<FocusIn>", lambda e: reparation_entry.delete('0', 'end'))
reparation_entry.grid(row=0, column=0, padx=20, pady=10, sticky="ew")

# Qte
qte_spinbox = ttk.Spinbox(widget2_frame, from_=1, to=100)
qte_spinbox.insert(0, "Qté")
qte_spinbox.grid(row=1, column=0, padx=20, pady=10, sticky="ew")

# Prix / unité H.T
Prix_entry = ttk.Entry(widget2_frame)
Prix_entry.insert(0, "Prix / unité H.T")
Prix_entry.bind("<FocusIn>", lambda e: Prix_entry.delete('0', 'end'))
Prix_entry.grid(row=2, column=0, padx=20, pady=10, sticky="ew")

# N° de facture
facture_entry = ttk.Entry(widget2_frame)
facture_entry.insert(0, "N° de facture")
facture_entry.bind("<FocusIn>", lambda e: facture_entry.delete('0', 'end'))
facture_entry.grid(row=3, column=0, padx=20, pady=10, sticky="ew")

# N° Carnert vignette
Ncarnet_entry = ttk.Entry(widget2_frame)
Ncarnet_entry.insert(0, "N° Carnert vignette")
Ncarnet_entry.bind("<FocusIn>", lambda e: Ncarnet_entry.delete('0', 'end'))
Ncarnet_entry.grid(row=4, column=0, padx=20, pady=10, sticky="ew")

# Button
Btn = ttk.Button(root, text="Insert", command=insert_row)
Btn.grid(row=2, column=0, padx=20, pady=10, sticky="ew")

# Separator
separator = ttk.Separator(root)
separator.grid(row=1, column=0, padx=20, pady=10, sticky="ew")

# Mode switch
mode_switch = ttk.Checkbutton(widget_frame, text="Mode", command=toggle_mode)
mode_switch.grid(row=4, column=0, sticky="nsew", padx=5, pady=10)

root.mainloop()
